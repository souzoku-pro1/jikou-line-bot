"""財産目録 xlsx の生成（ZAISAN-GEN-1・units 層）

- テンプレの正 = xlsx_templates/souzoku/財産目録.xlsx（大野書式の実物から
  make_zaisan_xlsx_template.py でデータ行をクリアして収載・SHA-256 pin＝
  契約書と同型。Ⅾ（支出）のローマ数字 U+216E 等の表記は原本のまま）
- 分類（A〜D）・特定情報パーサ・評価確定ガードは S3 unit（zaisan_mokuroku）
  を再利用し、描画層のみ xlsx（openpyxl 行差し込み）
- B 部の小計・総合計算定は全件「相続開始時残高」（大野裁定）。未入力行が
  あれば集計不能＝部分合計を出さず注記で明示（souzoku_dashboard の
  「0 円へ黙って落とさない」流儀）。「現在残高」は表示のみ
- 持分評価格 = 持分×評価額のサーバ側計算（「X分のY」/「Y/X」grammar のみ・
  端数切捨て）。grammar 外は空欄+備考に明示（field は設けない）
- 小計 4+総合計（= A+B+C−D・原本の式と同じ構成）はサーバ側 Python int
  集計。セル式は持たせない（verify が式の不在も検査）

ZAISAN-GEN-2（大野裁定・下書き生成）:
- 生成条件は「有効な財産レコード 1 件以上」のみ（評価確定=yes の全件要求を
  生成条件から外す）。ただし「機械は確定しない」は維持——評価確定≠yes の
  行は備考へ「※評価未確定」を自動追記し、未確定行が 1 つでもあれば
  ①表題直上の空き行（B1:K1 結合・原本の余白行）へ下書きバナーを表示
  ②算定できた小計・総合計の隣（J 列=注記欄）へ「（暫定）」を表示。
  金額セルは int のまま（暫定表示は注記欄に置き、検証水準を落とさない）。
- 全件確定+全金額入力なら従来どおり注記なしの完成版（現行検証水準のまま）。
- verify は下書きモードも対レコード照合（未確定注記・バナー・暫定表示を
  閉集合として照合）。
"""

import hashlib
import io
import re
from copy import copy
from datetime import date

from openpyxl import load_workbook

from config import get_office_info
from hub.docx_builder import to_wareki
from units.souzoku.zaisan_mokuroku import (
    ZaisanMokurokuError, _classify, _fudousan_row, _val, _yokin_row)

TEMPLATE_PATH = "xlsx_templates/souzoku/財産目録.xlsx"
# 収載現物（make_zaisan_xlsx_template.py 生成・コミット済み artifact）の pin。
# 再収載した場合は本値をスクリプトの出力で更新する
TEMPLATE_SHA256 = (
    "7dc01369a1da271ca99fc6a76726ad436977afd4402eb4197d82362af458a3a4")

# クリア済みテンプレの行配置（1-based・make_zaisan_xlsx_template.py と対）
ROW_A, SUB_A = 6, 7
ROW_B, SUB_B = 10, 11
ROW_C, SUB_C = 14, 15
ROW_D, SUB_D = 19, 20
ROW_TOTAL = 23
_COL_SUB = 9    # I 列（小計・総合計）
_COL_NOTE = 10  # J 列（備考・集計不能注記）

# 集計不能注記（部分合計を出さない＝0 円へ黙って落とさない）
NOTE_A = "評価額・持分の不足により小計は算定していません"
NOTE_B = "相続開始時残高が未入力の行があるため小計は算定していません"
NOTE_CD = "金額が未入力の行があるため小計は算定していません"
NOTE_TOTAL = "小計に算定不能があるため総合計は算定していません"
NOTE_MOCHIBUN = "持分評価格は自動算定できません（持分の書式外）"
# ZAISAN-GEN-2: 下書きモードの明示（弁護士確定は人の行為のまま＝機械は
# 確定しない。確定版と紛れない見た目を注記欄・余白行で実現し体裁は不変）
NOTE_UNCONFIRMED = "※評価未確定"
DRAFT_BANNER = "【下書き・評価未確定の項目があります】"
PROVISIONAL_MARK = "（暫定）"


def _is_unconfirmed(record: dict) -> bool:
    return _val(record, "評価確定") != "yes"


def count_unconfirmed(records: list[dict]) -> int:
    """評価確定≠yes の行数（webhook の生成通知が下書き/完成版の明記に使う）。"""
    return sum(1 for r in records if _is_unconfirmed(r))


def _note_with_flags(record: dict, note: str) -> str:
    """行の備考へ未確定注記を自動追記（GEN-2・既存注記とは「／」連結）。"""
    if _is_unconfirmed(record):
        note = f"{note}／{NOTE_UNCONFIRMED}" if note else NOTE_UNCONFIRMED
    return note

_MOCHIBUN_KANJI = re.compile(r"(\d+)分の(\d+)")
_MOCHIBUN_SLASH = re.compile(r"(\d+)/(\d+)")


class ZaisanXlsxIntegrityError(RuntimeError):
    """テンプレ pin 不一致・生成物の検証失敗（添付しない）。"""


def verify_template_integrity() -> None:
    data = open(TEMPLATE_PATH, "rb").read()
    if hashlib.sha256(data).hexdigest() != TEMPLATE_SHA256:
        raise ZaisanXlsxIntegrityError("template hash mismatch")


def _mochibun_fraction(text: str) -> tuple[int, int] | None:
    """持分文字列 →（分子, 分母）。「2分の1」=1/2・「1/2」のみ。

    fix1[03]: 範囲固定 0 < 分子 <= 分母 を必須化。範囲外（1分の2・3分の4
    等）・0 分母・0 分子は None（空欄+固定注記 NOTE_MOCHIBUN の経路）。
    持分評価格の端数は**切捨て**（builder の // 演算・テストで固定）。"""
    t = text.replace("　", "").replace(" ", "")
    m = _MOCHIBUN_KANJI.fullmatch(t)
    if m:
        num, den = int(m.group(2)), int(m.group(1))
    else:
        m = _MOCHIBUN_SLASH.fullmatch(t)
        if not m:
            return None
        num, den = int(m.group(1)), int(m.group(2))
    if not 0 < num <= den:
        return None
    return num, den


_AMOUNT_INT_RE = re.compile(r"[0-9]+")


def _int_or_none(record: dict, code: str) -> int | None:
    """fix1[01]: 金額の受理を閉集合化——bool を除く実 int、または 10 進整数
    文字列（^[0-9]+$・既存金額 grammar と同基準）のみ。小数・指数表記・
    bool・桁区切り・符号は不受理=None（当該行は集計不能の既存流儀へ）。
    int(float(...)) 経由の精度喪失（2^53 超）も本閉集合で排除される。"""
    raw = (record.get(code) or {}).get("value")
    if isinstance(raw, bool):
        return None
    if isinstance(raw, int):
        return raw
    if isinstance(raw, str) and _AMOUNT_INT_RE.fullmatch(raw):
        return int(raw)
    return None


def _a_row(record: dict) -> tuple[dict, int | None]:
    """A 部（不動産）。effective = 持分評価格（持分あり）/評価額（持分なし）。"""
    base = _fudousan_row(record)  # 特定情報 slash-KV（既存パーサ）
    hyoka = _int_or_none(record, "評価額")
    note = _val(record, "備考")
    share = None
    if base["持分"]:
        frac = _mochibun_fraction(base["持分"])
        if frac is None:
            note = f"{note}／{NOTE_MOCHIBUN}" if note else NOTE_MOCHIBUN
            effective = None
        elif hyoka is None:
            effective = None
        else:
            share = hyoka * frac[0] // frac[1]  # 端数切捨て
            effective = share
    else:
        effective = hyoka
    cells = {3: base["所在"], 4: base["地番家屋番号"], 5: base["地目種別"],
             6: base["地積床面積"], 7: base["持分"], 8: hyoka, 9: share,
             10: _note_with_flags(record, note),
             11: _val(record, "資料番号")}
    return cells, effective


def _b_row(record: dict) -> tuple[dict, int | None]:
    """B 部（預貯金）。effective = 相続開始時残高（大野裁定・現在残高は表示のみ）。"""
    base = _yokin_row(record)  # 特定情報トークン（既存パーサ）
    bank = base["金融機関"]
    if base["支店"]:
        bank = f"{bank}　{base['支店']}"
    start = _int_or_none(record, "相続開始時残高")
    now = _int_or_none(record, "現在残高")
    cells = {3: bank, 4: base["種別"], 5: base["口座番号"], 8: start, 9: now,
             10: _note_with_flags(record, _val(record, "備考")),
             11: _val(record, "資料番号")}
    return cells, start


def _cd_row(record: dict, company_col: int) -> tuple[dict, int | None]:
    """C/D 部。種類=財産種別・会社名等=特定情報（全文＝情報を落とさない）。
    会社名等の列は C 部=F 列（6）・D 部=D 列（4）と異なる（原本の列構成）。"""
    amount = _int_or_none(record, "評価額")
    cells = {3: _val(record, "財産種別"),
             company_col: _val(record, "特定情報") or None,
             9: amount,
             10: _note_with_flags(record, _val(record, "備考")),
             11: _val(record, "資料番号")}
    return cells, amount


def _subtotal(effectives: list[int | None]) -> int | None:
    """1 件でも算定不能があれば None（部分合計を出さない）。空リストは 0。"""
    if any(e is None for e in effectives):
        return None
    return sum(effectives)


def _expand(ws, proto_row: int, n: int) -> None:
    """プロトタイプ行の下に n-1 行を挿入し、書式・行高を追随させる。

    注意: openpyxl の insert_rows は結合レンジを移動させない（delete_rows と
    同様）。結合は build 側で「全解除 → 挿入・差し込み → 最終座標で一括
    再構築（_apply_merges）」する方式で扱う。"""
    if n <= 1:
        return
    ws.insert_rows(proto_row + 1, n - 1)
    for i in range(1, n):
        r = proto_row + i
        ws.row_dimensions[r].height = ws.row_dimensions[proto_row].height
        for c in range(2, 12):
            ws.cell(row=r, column=c)._style = copy(
                ws.cell(row=proto_row, column=c)._style)


def _merge_ranges(na: int, nb: int, nc: int, nd: int) -> list[str]:
    """可変行数（各部の描画行数・0 件の部は 1）から最終結合レンジを決定的に
    算出する（テンプレの正準レンジ+データ行の 3 結合パターン: B=口座番号
    E:G・C/D=種別/会社名等 D:E と F:H）。builder の再構築と fix1[04] の
    生成物検証（完全一致検査）が共用する単一の正。"""
    oa, ob, oc, od = na - 1, nb - 1, nc - 1, nd - 1
    merges = ["B1:K1", "B2:I2", "B4:K4"]
    row_b_label = 8 + oa
    merges += [f"B{row_b_label}:K{row_b_label}",
               f"E{row_b_label + 1}:G{row_b_label + 1}"]
    for i in range(nb):
        r = ROW_B + oa + i
        merges.append(f"E{r}:G{r}")
    row_c_label = 12 + oa + ob
    merges += [f"B{row_c_label}:K{row_c_label}",
               f"D{row_c_label + 1}:E{row_c_label + 1}",
               f"F{row_c_label + 1}:H{row_c_label + 1}"]
    for i in range(nc):
        r = ROW_C + oa + ob + i
        merges += [f"D{r}:E{r}", f"F{r}:H{r}"]
    row_d_label = 17 + oa + ob + oc
    merges += [f"B{row_d_label}:K{row_d_label}",
               f"D{row_d_label + 1}:E{row_d_label + 1}",
               f"F{row_d_label + 1}:H{row_d_label + 1}"]
    for i in range(nd):
        r = ROW_D + oa + ob + oc + i
        merges += [f"D{r}:E{r}", f"F{r}:H{r}"]
    total_row = ROW_TOTAL + oa + ob + oc + od
    merges.append(f"J{total_row}:K{total_row}")
    return merges


def _apply_merges(ws, na: int, nb: int, nc: int, nd: int) -> None:
    for rng in _merge_ranges(na, nb, nc, nd):
        ws.merge_cells(rng)


def _write_section(ws, proto_row: int, rows: list[dict]) -> None:
    for i, cells in enumerate(rows):
        r = proto_row + i
        ws.cell(row=r, column=2).value = i + 1  # 番号
        for c, v in cells.items():
            if v not in (None, ""):
                ws.cell(row=r, column=c).value = v


def build_zaisan_xlsx(records: list[dict], *,
                      decedent_name: str | None = None,
                      created: date | None = None) -> bytes:
    """財産行（App 財産のレコード）から財産目録 xlsx を組み立てる
    （kintone I/O なし・呼び出し側は verify_zaisan_xlsx を通してから添付）。"""
    if not records:
        raise ZaisanMokurokuError(
            "財産行が0件です（App 財産に案件の財産が登録されていません）")
    # ZAISAN-GEN-2 裁定: 生成条件は「有効な財産行 1 件以上」のみ
    # （評価確定の全件要求は生成条件から外し、下書きモードで明示する）
    verify_template_integrity()

    fud, yok, son, sai = _classify(records)
    a = [_a_row(r) for r in fud]
    b = [_b_row(r) for r in yok]
    c = [_cd_row(r, 6) for r in son]
    d = [_cd_row(r, 4) for r in sai]

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # 結合を全解除（openpyxl の insert_rows は結合を移動させないため、
    # 挿入・差し込み後に _apply_merges で最終座標へ一括再構築する）
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))

    # 下のセクションから挿入（上のセクションの行番号を動かさない）
    _expand(ws, ROW_D, len(d))
    _expand(ws, ROW_C, len(c))
    _expand(ws, ROW_B, len(b))
    _expand(ws, ROW_A, len(a))

    oa, ob, oc = (max(len(x), 1) - 1 for x in (a, b, c))
    od = max(len(d), 1) - 1
    _write_section(ws, ROW_A, [x[0] for x in a])
    _write_section(ws, ROW_B + oa, [x[0] for x in b])
    _write_section(ws, ROW_C + oa + ob, [x[0] for x in c])
    _write_section(ws, ROW_D + oa + ob + oc, [x[0] for x in d])

    # ZAISAN-GEN-2: 未確定行が 1 つでもあれば下書きモード——表題直上の
    # 余白行（B1:K1 結合・原本の空行）へバナー、算定できた小計・総合計の
    # 注記欄（J 列）へ「（暫定）」。金額セルは int のまま（検証水準不変）
    draft = any(_is_unconfirmed(r) for r in records)
    if draft:
        ws.cell(row=1, column=2).value = DRAFT_BANNER
    subs = {}
    for key, rows, sub_row, note in (
            ("A", a, SUB_A + oa, NOTE_A),
            ("B", b, SUB_B + oa + ob, NOTE_B),
            ("C", c, SUB_C + oa + ob + oc, NOTE_CD),
            ("D", d, SUB_D + oa + ob + oc + od, NOTE_CD)):
        sub = _subtotal([x[1] for x in rows])
        subs[key] = sub
        if sub is None:
            ws.cell(row=sub_row, column=_COL_NOTE).value = note
        else:
            ws.cell(row=sub_row, column=_COL_SUB).value = sub
            if draft:
                ws.cell(row=sub_row, column=_COL_NOTE).value =                     PROVISIONAL_MARK
    total_row = ROW_TOTAL + oa + ob + oc + od
    if any(v is None for v in subs.values()):
        ws.cell(row=total_row, column=_COL_NOTE).value = NOTE_TOTAL
    else:
        ws.cell(row=total_row, column=_COL_SUB).value = (
            subs["A"] + subs["B"] + subs["C"] - subs["D"])
        if draft:
            ws.cell(row=total_row, column=_COL_NOTE).value = PROVISIONAL_MARK

    # 見出し差し込み（テンプレの {{...}} を置換）
    name = decedent_name or next(
        (_val(r, "被相続人名表示用") for r in records
         if _val(r, "被相続人名表示用")), "")
    office = get_office_info()
    author = "　".join(x for x in (
        office["名称"],
        f"弁護士　{office['弁護士名']}" if office["弁護士名"] else "",
    ) if x)
    reps = {"{{被相続人名}}": name,
            "{{作成日}}": to_wareki(created or date.today()),
            "{{作成者}}": author}
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if isinstance(cell.value, str) and "{{" in cell.value:
                v = cell.value
                for k, val in reps.items():
                    v = v.replace(k, val)
                cell.value = v

    _apply_merges(ws, max(len(a), 1), max(len(b), 1),
                  max(len(c), 1), max(len(d), 1))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _amount_cell_ok(v) -> bool:
    return v is None or (isinstance(v, int) and not isinstance(v, bool))


def verify_zaisan_xlsx(xlsx_bytes: bytes, records: list[dict]) -> None:
    """生成物の添付前検証（ZAISAN-GEN-1 要件 3+fix1[02][04]）。

    - 行数整合: 各部の描画行数 = レコード数（0 件の部は空 1 行）・総行数一致
    - 金額セルは int のみ（float/str/bool を拒否）・セル式（"=..."）の不在
    - fix1[02] 対レコード照合: 元 records から期待する閉集合投影
      （部区分・番号・名称/特定情報・金額・持分/持分評価格・備考・資料番号）
      を再構成し、明細行ごとにセル単位で完全一致を検査。小計/総合計も投影
      からの再計算と照合（集計不能の部は小計空欄+注記の存在を要求・
      総合計 = A+B+C−D）。明細と集計を整合的に同時改変した生成物も、
      元データとの不一致として拒否される
    - fix1[04] レイアウト: 可変行数から決定的に算出した最終結合レンジ集合
      （_merge_ranges・builder と単一の正）と merged_cells の完全一致
    - GEN-2 下書きモード: 未確定注記（備考・行投影に含む）・下書きバナー
      （B1）・暫定表示（小計/総合計の J 列）も閉集合として照合——完成版に
      バナー/暫定が付くこと・下書きに欠けることの両方向を拒否
    - GEN2-fix1: 集計不能時の注記欄は部ごとの期待注記（A=NOTE_A/B=NOTE_B/
      C・D=NOTE_CD/総合計=NOTE_TOTAL）と完全一致（draft/final とも不変＝
      集計不能注記が暫定表示に優先。PROVISIONAL_MARK や任意文字列は拒否）
    """
    fud, yok, son, sai = _classify(records)
    a = [_a_row(r) for r in fud]
    b = [_b_row(r) for r in yok]
    c = [_cd_row(r, 6) for r in son]
    d = [_cd_row(r, 4) for r in sai]
    na, nb, nc, nd = (max(len(x), 1) for x in (a, b, c, d))
    oa, ob, oc = na - 1, nb - 1, nc - 1
    od = nd - 1
    ws = load_workbook(io.BytesIO(xlsx_bytes)).active
    if ws.max_row != ROW_TOTAL + oa + ob + oc + od:
        raise ZaisanXlsxIntegrityError("row count mismatch")
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                raise ZaisanXlsxIntegrityError("formula found")
    got_merges = {str(r) for r in ws.merged_cells.ranges}
    if got_merges != set(_merge_ranges(na, nb, nc, nd)):
        raise ZaisanXlsxIntegrityError("merge ranges mismatch")
    draft = any(_is_unconfirmed(r) for r in records)
    banner = ws.cell(row=1, column=2).value
    if banner != (DRAFT_BANNER if draft else None):
        raise ZaisanXlsxIntegrityError("draft banner mismatch")

    sections = (
        ("A", ROW_A, a, (8, 9)),
        ("B", ROW_B + oa, b, (8, 9)),
        ("C", ROW_C + oa + ob, c, (9,)),
        ("D", ROW_D + oa + ob + oc, d, (9,)),
    )
    subs = {}
    for key, start, rows, amount_cols in sections:
        for i in range(max(len(rows), 1)):
            r = start + i
            for col in amount_cols:
                if not _amount_cell_ok(ws.cell(row=r, column=col).value):
                    raise ZaisanXlsxIntegrityError("non-int amount cell")
            if i < len(rows):
                if ws.cell(row=r, column=2).value != i + 1:
                    raise ZaisanXlsxIntegrityError("番号 sequence mismatch")
                cells = rows[i][0]
                for col in range(3, 12):
                    want = cells.get(col)
                    if want in (None, ""):
                        want = None
                    if ws.cell(row=r, column=col).value != want:
                        raise ZaisanXlsxIntegrityError(
                            f"cell mismatch section={key} "
                            f"row={i + 1} col={col}")
            else:  # 空セクションのプレースホルダ行は完全空欄
                for col in range(2, 12):
                    if ws.cell(row=r, column=col).value is not None:
                        raise ZaisanXlsxIntegrityError(
                            f"unexpected value in empty section {key}")
        subs[key] = _subtotal([x[1] for x in rows])

    # GEN2-fix1（ZAISAN-GEN2-01）: 集計不能注記は部ごとの期待注記との
    # **完全一致**で照合（非空存在検査を廃止）。期待値は draft/final で
    # 不変＝集計不能注記が PROVISIONAL_MARK に優先する
    expected_notes = {"A": NOTE_A, "B": NOTE_B, "C": NOTE_CD, "D": NOTE_CD}
    for key, sub_row in (("A", SUB_A + oa), ("B", SUB_B + oa + ob),
                         ("C", SUB_C + oa + ob + oc),
                         ("D", SUB_D + oa + ob + oc + od)):
        got = ws.cell(row=sub_row, column=_COL_SUB).value
        note = ws.cell(row=sub_row, column=_COL_NOTE).value
        if subs[key] is None:
            if got is not None or note != expected_notes[key]:
                raise ZaisanXlsxIntegrityError(
                    f"subtotal {key}: uncomputable note mismatch")
        else:
            if got != subs[key]:
                raise ZaisanXlsxIntegrityError(f"subtotal {key} mismatch")
            # GEN-2: 暫定表示の閉集合照合（下書き=（暫定）・完成版=空）
            if note != (PROVISIONAL_MARK if draft else None):
                raise ZaisanXlsxIntegrityError(
                    f"subtotal {key}: provisional mark mismatch")
    total_row = ROW_TOTAL + oa + ob + oc + od
    got_total = ws.cell(row=total_row, column=_COL_SUB).value
    total_note = ws.cell(row=total_row, column=_COL_NOTE).value
    if any(v is None for v in subs.values()):
        if got_total is not None or total_note != NOTE_TOTAL:
            raise ZaisanXlsxIntegrityError(
                "total: uncomputable note mismatch")
    else:
        if got_total != subs["A"] + subs["B"] + subs["C"] - subs["D"]:
            raise ZaisanXlsxIntegrityError("total mismatch")
        if total_note != (PROVISIONAL_MARK if draft else None):
            raise ZaisanXlsxIntegrityError("total: provisional mark mismatch")
