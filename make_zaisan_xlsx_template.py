"""財産目録 xlsx テンプレートの収載スクリプト（ZAISAN-GEN-1）

大野保存の実物（Desktop/claude/財産目録　五十嵐資幸.xlsx・書式の正）から、
実案件データをクリアしたテンプレートを xlsx_templates/souzoku/財産目録.xlsx
へ生成する。契約書（CONTRACT-GEN-1）と同型で、生成物の SHA-256 を
units/souzoku/zaisan_xlsx.py の TEMPLATE_SHA256 に pin する。

クリア内容（構造・列幅・罫線・金額書式・行高・Ⅾ（支出）等の表記は原本のまま）:
  - 見出し差し込み: C3=被相続人　{{被相続人名}} / J2={{作成日}} / J3={{作成者}}
  - 各部のデータ行を 1 行（プロトタイプ行・値なし）に縮約
  - 小計 4+総合計のセル式を除去（生成時にサーバ側 int 集計で書く）
  - 末尾の空行（r42 以降）を削除
  - 結合セルは正準 17 レンジへ決定的に再構築（openpyxl の行削除に伴う
    結合ズレを排除し、テンプレ構造をテストで pin 可能にする）

実行: python make_zaisan_xlsx_template.py
"""

import hashlib
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path.home() / "Desktop" / "claude" / "財産目録　五十嵐資幸.xlsx"
OUT = Path("xlsx_templates") / "souzoku" / "財産目録.xlsx"

# 原本の行配置（1-based）。データ行はプロトタイプ 1 行を残して削除する
_DELETES = [  # 下から実行（行番号を動かさないため）
    (42, 15),  # 末尾の空行 r42-r56
    (34, 4),   # D 部データ r34-r37（r33 を残す）
    (24, 5),   # C 部データ r24-r28（r23 を残す）
    (13, 7),   # B 部データ r13-r19（r12 を残す）
    (7, 2),    # A 部データ r7-r8（r6 を残す）
]
# 削除後に残る行の原本行番号（行高の再配置に使う）
_KEPT = [1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 20, 21, 22, 23,
         29, 30, 31, 32, 33, 38, 39, 40, 41]
# クリア後の正準結合レンジ（構造 pin・test_zaisan_gen1 が同一集合を検査）
CANON_MERGES = [
    "B1:K1", "B2:I2", "B4:K4",           # 余白・表題・A 部見出し
    "B8:K8", "E9:G9", "E10:G10",         # B 部見出し・列頭・プロト行
    "B12:K12", "D13:E13", "F13:H13",     # C 部見出し・列頭
    "D14:E14", "F14:H14",                # C 部プロト行
    "B17:K17", "D18:E18", "F18:H18",     # D 部見出し・列頭
    "D19:E19", "F19:H19",                # D 部プロト行
    "J23:K23",                           # 総合計行の注記欄
]
# クリア対象（クリア後座標）: プロトタイプ行の値・小計/総合計セル
_PROTO_ROWS = (6, 10, 14, 19)
_VALUE_CELLS = ("I7", "I11", "I15", "I20", "I23")


def main() -> None:
    if not SOURCE.exists():
        raise SystemExit(f"書式の正が見つかりません: {SOURCE}")
    wb = load_workbook(SOURCE)
    ws = wb.active

    # 行高を退避（openpyxl の delete_rows は row_dimensions を移動しない）
    heights = {old: ws.row_dimensions[old].height
               for old in _KEPT if old in ws.row_dimensions}

    # 見出しの差し込み化（値のみ・書式は原本のまま）
    ws["C3"] = "被相続人　{{被相続人名}}"
    ws["J2"] = "{{作成日}}"
    ws["J3"] = "{{作成者}}"

    # 結合をすべて解除してから行削除（openpyxl の delete_rows は結合レンジを
    # 追随させないため、削除後に正準集合を組み直す）
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))

    for start, amount in _DELETES:
        ws.delete_rows(start, amount)

    # 行高の再配置
    for new, old in enumerate(_KEPT, start=1):
        if old in heights and heights[old] is not None:
            ws.row_dimensions[new].height = heights[old]
    for idx in list(ws.row_dimensions):
        if idx > len(_KEPT):
            del ws.row_dimensions[idx]

    # プロトタイプ行・集計セルの値クリア（書式は保持される・結合再構築の前＝
    # MergedCell 化されると anchor 以外へ書けないため）
    for r in _PROTO_ROWS:
        for c in range(2, 12):
            ws.cell(row=r, column=c).value = None
    for cell in _VALUE_CELLS:
        ws[cell].value = None

    # 結合セルを正準集合へ決定的に再構築
    for rng in CANON_MERGES:
        ws.merge_cells(rng)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    sha = hashlib.sha256(OUT.read_bytes()).hexdigest()
    print(f"saved: {OUT} rows={ws.max_row}")
    print(f"TEMPLATE_SHA256 = {sha}")


if __name__ == "__main__":
    main()
