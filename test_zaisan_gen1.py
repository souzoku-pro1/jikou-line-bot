"""ZAISAN-GEN-1: 財産目録 xlsx の自動生成。

固定する仕様:
- テンプレの正 = xlsx_templates/souzoku/財産目録.xlsx（大野書式の収載現物・
  SHA-256 pin・23 行・正準 17 結合・Ⅾ（支出）=U+216E 等の表記維持）
- 分類 A〜D・特定情報パーサ・評価確定ガード（確定は弁護士・不変）は
  S3 unit を再利用。描画層のみ openpyxl 行差し込み（結合 3 パターン追随）
- B 部の小計・総合計算定は全件「相続開始時残高」（大野裁定・現在残高は
  表示のみ）。未入力行があれば集計不能＝部分合計を出さず注記（0 円へ
  黙って落とさない）
- 持分評価格 = 持分×評価額のサーバ側計算（X分のY / Y/X のみ・端数切捨て・
  grammar 外は空欄+備考明示）
- 小計 4+総合計（= A+B+C−D）はサーバ側 int 集計・セル式は持たせない
- 生成 xlsx の添付前検証: 金額セル int・小計/総合計の再計算一致・行数整合・
  セル式不在・番号連番（verify_zaisan_xlsx）
- webhook は CONTRACT-GEN 確立構造の同型（App26・入口ガード・CAS 4 値・
  409 のみ cas_lost・fail-closed=財産行 0 件/評価未確定は状態不変+通知）

fix1（R-ZAISAN-GEN-1）:
- [01] 金額受理の閉集合化: bool を除く実 int と 10 進整数文字列（^[0-9]+$）
  のみ。小数・指数・bool・桁区切り・符号・全角は不受理=当該行を集計不能
  扱い（int(float(...)) の精度喪失も排除・"9007199254740993" が正確値）
- [02] 添付前検証の対レコード照合: 元 records の閉集合投影と明細セル単位で
  完全一致（明細+小計+総合計の整合的同時改変も拒否）
- [03] 持分 grammar の範囲固定: 0 < 分子 <= 分母 必須（1分の2・3分の4 等は
  空欄+固定注記）・端数切捨てを pin
- [04] レイアウト検証: 可変行数から決定的算出した結合レンジ集合と
  merged_cells の完全一致（0 件/1 件/複数/多数行の各配置・書式保持）
"""

import hashlib
import io
import os
import unittest
from unittest.mock import AsyncMock, patch

_ENV = {
    "ANTHROPIC_API_KEY": "dummy", "LINE_CHANNEL_SECRET": "dummy_secret",
    "LINE_CHANNEL_ACCESS_TOKEN": "dummy_token", "KINTONE_SUBDOMAIN": "testsub",
    "KINTONE_APP_ID": "21", "KINTONE_API_TOKEN": "dummy",
    "SOUZOKU_KINTONE_APP_ID": "26", "SOUZOKU_KINTONE_API_TOKEN": "dummy",
    "CLOUDSIGN_CLIENT_ID": "c", "CLOUDSIGN_WEBHOOK_SECRET": "cs",
    "KINTONE_WEBHOOK_TOKEN": "kintone-token",
    "DOCUMENT_WEBHOOK_SECRET": "doc-secret",
    "APP_APPROVAL": "29", "TOKEN_APPROVAL": "d", "HEALTHCHECK_DISABLED": "1",
    "STRIPE_WEBHOOK_SECRET": "w", "GOOGLE_VISION_API_KEY": "dummy_vision",
    "APP_ZAISAN": "35", "TOKEN_ZAISAN": "dummy",
    "OFFICE_NAME": "大野法律事務所", "OFFICE_ATTORNEY": "大野太郎",
}
for _k, _v in _ENV.items():
    os.environ.setdefault(_k, _v)

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

import main  # noqa: E402
import zaisan_webhook as zw  # noqa: E402
from config import EXPECTED_KINTONE_SCHEMA  # noqa: E402
from hub.kintone import KintoneError  # noqa: E402
from make_zaisan_xlsx_template import CANON_MERGES  # noqa: E402
from units.souzoku import zaisan_xlsx as zx  # noqa: E402
from units.souzoku.zaisan_mokuroku import ZaisanMokurokuError  # noqa: E402

_client = TestClient(main.app)
_URL = "/zaisan/doc-secret"


def _zrec(record_id="1", **over):
    base = {"$id": record_id, "財産種別": "預貯金", "特定情報": "",
            "名義": "", "評価額": "", "評価確定": "yes", "備考": "",
            "資料番号": "", "相続開始時残高": "", "現在残高": "",
            "被相続人名表示用": "試験太郎", "有効": "yes"}
    base.update(over)
    return {k: {"value": v} for k, v in base.items()}


def _sample_records():
    return [
        _zrec("1", 財産種別="不動産_土地",
              特定情報="所在 川口市青木1丁目 / 地番 12番3 / 地目 宅地 / "
                       "地積 120.5㎡ / 持分 2分の1",
              評価額="1000001", 資料番号="A-1"),
        _zrec("2", 財産種別="不動産_建物",
              特定情報="所在 川口市青木1丁目 / 家屋番号 12番3 / 種類 居宅 / "
                       "床面積 80㎡",
              評価額="2000000"),
        _zrec("3", 財産種別="預貯金",
              特定情報="テスト銀行 青木支店 普通預金 口座番号1234567",
              相続開始時残高="300000", 現在残高="299000", 資料番号="B-1"),
        _zrec("4", 財産種別="預貯金",
              特定情報="サンプル銀行 本店 定期預金 口座番号7654321",
              相続開始時残高="100000"),
        _zrec("5", 財産種別="有価証券", 特定情報="テスト証券の株式",
              評価額="400000"),
        _zrec("6", 財産種別="葬儀費用", 特定情報="テスト葬祭",
              評価額="250000", 備考="立替済"),
    ]
# 期待値: A小計=1000001//2 + 2000000 = 2500000 / B=400000 / C=400000 /
# D=250000 / 総合計 = A+B+C−D = 3050000


class TestTemplatePin(unittest.TestCase):
    def test_sha256_pinned(self):
        data = open(zx.TEMPLATE_PATH, "rb").read()
        self.assertEqual(hashlib.sha256(data).hexdigest(), zx.TEMPLATE_SHA256)

    def test_structure(self):
        ws = load_workbook(zx.TEMPLATE_PATH).active
        self.assertEqual(ws.max_row, 23)
        self.assertEqual(ws["B4"].value, "Ａ(不動産)")
        self.assertIn("預貯金", ws["B8"].value)
        self.assertEqual(ws["B17"].value, "Ⅾ（支出）")  # ローマ数字維持
        self.assertEqual(ws["H9"].value, "相続開始時残高(円)")
        self.assertEqual(ws["I9"].value, "現在残高(円)")
        self.assertEqual(ws["I5"].value, "持分評価格")
        self.assertIn("{{被相続人名}}", ws["C3"].value)
        self.assertEqual(ws["J2"].value, "{{作成日}}")
        self.assertEqual(ws["J3"].value, "{{作成者}}")
        self.assertEqual(sorted(str(r) for r in ws.merged_cells.ranges),
                         sorted(CANON_MERGES))
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    self.assertFalse(cell.value.startswith("="))
        # 金額書式（原本の #,##0）がプロト行・小計セルに残っている
        for ref in ("H6", "I6", "H10", "I10", "I14", "I19",
                    "I7", "I11", "I15", "I20", "I23"):
            self.assertIn("#,##0", ws[ref].number_format)


class TestMochibunGrammar(unittest.TestCase):
    def test_closed_grammar(self):
        cases = {
            "2分の1": (1, 2), "3分の2": (2, 3), "1/2": (1, 2),
            "10分の3": (3, 10), "２分の１くらい": None, "持分不明": None,
            "0分の1": None, "": None, "2分の1くらい": None,
            # fix1[03]: 範囲固定 0 < 分子 <= 分母（範囲外・0 分子は不受理）
            "1分の2": None, "3分の4": None, "2/1": None, "0/2": None,
            "1分の0": None, "2分の2": (2, 2),
        }
        for text, want in cases.items():
            with self.subTest(text=text):
                self.assertEqual(zx._mochibun_fraction(text), want)

    def test_floor_division_pinned(self):
        # fix1[03]: 端数は切捨て（100×1/3 = 33・仕様固定）
        records = _sample_records()
        records[0]["特定情報"] = {"value":
                                  "所在 川口市青木1丁目 / 持分 3分の1"}
        records[0]["評価額"] = {"value": "100"}
        with patch.dict(os.environ, _ENV):
            data = zx.build_zaisan_xlsx(records)
        ws = load_workbook(io.BytesIO(data)).active
        self.assertEqual(ws["I6"].value, 33)
        zx.verify_zaisan_xlsx(data, records)

    def test_out_of_range_mochibun_blank_plus_note(self):
        # fix1[03]: 範囲外（1分の2）は空欄+固定注記
        records = _sample_records()
        records[0]["特定情報"] = {"value":
                                  "所在 川口市青木1丁目 / 持分 1分の2"}
        with patch.dict(os.environ, _ENV):
            data = zx.build_zaisan_xlsx(records)
        ws = load_workbook(io.BytesIO(data)).active
        self.assertIsNone(ws["I6"].value)
        self.assertIn(zx.NOTE_MOCHIBUN, ws["J6"].value)
        self.assertEqual(ws["J8"].value, zx.NOTE_A)
        zx.verify_zaisan_xlsx(data, records)


class TestAmountGrammar(unittest.TestCase):
    """fix1[01]: 金額受理の閉集合（既存金額 grammar と同基準）。"""

    def _v(self, raw):
        return zx._int_or_none({"金額": {"value": raw}}, "金額")

    def test_accepted_forms(self):
        self.assertEqual(self._v("100"), 100)
        self.assertEqual(self._v(100), 100)
        self.assertEqual(self._v("0"), 0)
        # 精度喪失検知: 2^53+1 が正確値のまま（int(float()) だと ...992 に化ける）
        self.assertEqual(self._v("9007199254740993"), 9007199254740993)
        self.assertNotEqual(self._v("9007199254740993"), 9007199254740992)

    def test_rejected_forms(self):
        for raw in ("100.9", 1.9, True, False, "1e6", "1E6", "1,000",
                    "-5", "+5", " 100", "100 ", "", None, "１００"):
            with self.subTest(raw=repr(raw)):
                self.assertIsNone(self._v(raw))

    def test_decimal_amount_uncomputable_flow(self):
        # 小数の評価額は正常値へ変換せず当該部を集計不能（注記連鎖）へ
        records = _sample_records()
        records[4]["評価額"] = {"value": "100.9"}
        with patch.dict(os.environ, _ENV):
            data = zx.build_zaisan_xlsx(records)
        ws = load_workbook(io.BytesIO(data)).active
        self.assertIsNone(ws["I16"].value)        # C 明細は空欄
        self.assertIsNone(ws["I17"].value)        # C 小計は出さない
        self.assertEqual(ws["J17"].value, zx.NOTE_CD)
        self.assertIsNone(ws["I25"].value)        # 総合計も出さない
        zx.verify_zaisan_xlsx(data, records)


class TestBuilder(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        with patch.dict(os.environ, _ENV):
            cls.data = zx.build_zaisan_xlsx(_sample_records())
        cls.ws = load_workbook(io.BytesIO(cls.data)).active

    def test_verify_passes_and_zip_header(self):
        self.assertTrue(self.data.startswith(b"PK"))
        zx.verify_zaisan_xlsx(self.data, _sample_records())

    def test_layout_and_amounts(self):
        ws = self.ws
        # A 部（2 行・oa=1）: 行 6-7・小計 8
        self.assertEqual(ws["B6"].value, 1)
        self.assertEqual(ws["C6"].value, "川口市青木1丁目")
        self.assertEqual(ws["G6"].value, "2分の1")
        self.assertEqual(ws["H6"].value, 1000001)
        self.assertEqual(ws["I6"].value, 500000)      # 端数切捨て
        self.assertEqual(ws["B7"].value, 2)
        self.assertIsNone(ws["I7"].value)             # 持分なし＝空欄
        self.assertEqual(ws["I8"].value, 2500000)
        self.assertEqual(ws["K6"].value, "A-1")       # 資料番号
        # B 部（2 行・ob=1）: 行 11-12・小計 13
        self.assertEqual(ws["C11"].value, "テスト銀行　青木支店")
        self.assertEqual(ws["D11"].value, "普通預金")
        self.assertEqual(ws["E11"].value, "1234567")
        self.assertEqual(ws["H11"].value, 300000)
        self.assertEqual(ws["I11"].value, 299000)     # 現在残高は表示のみ
        self.assertEqual(ws["H12"].value, 100000)
        self.assertEqual(ws["I13"].value, 400000)     # Σ相続開始時残高
        # C 部（1 行）: 行 16・小計 17
        self.assertEqual(ws["C16"].value, "有価証券")
        self.assertEqual(ws["F16"].value, "テスト証券の株式")
        self.assertEqual(ws["I16"].value, 400000)
        self.assertEqual(ws["I17"].value, 400000)
        # D 部（1 行）: 行 21・小計 22
        self.assertEqual(ws["C21"].value, "葬儀費用")
        self.assertEqual(ws["D21"].value, "テスト葬祭")
        self.assertEqual(ws["I21"].value, 250000)
        self.assertEqual(ws["J21"].value, "立替済")
        self.assertEqual(ws["I22"].value, 250000)
        # 総合計 = A+B+C−D（行 25）・総行数
        self.assertEqual(ws["I25"].value, 3050000)
        self.assertEqual(ws.max_row, 25)

    def test_inserted_rows_keep_merges(self):
        merges = {str(r) for r in self.ws.merged_cells.ranges}
        self.assertIn("E11:G11", merges)   # B プロト行（口座番号）
        self.assertIn("E12:G12", merges)   # B 挿入行に追随
        self.assertIn("D21:E21", merges)   # D 行（会社名等）
        self.assertIn("F21:H21", merges)

    def test_header_scalars(self):
        self.assertEqual(self.ws["C3"].value, "被相続人　試験太郎")
        self.assertIn("年", self.ws["J2"].value)      # 和暦作成日
        self.assertIn("大野太郎", self.ws["J3"].value)
        self.assertNotIn("{{", str(self.ws["C3"].value))

    def test_no_formulas_anywhere(self):
        for row in self.ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    self.assertFalse(cell.value.startswith("="))


class TestFailClosed(unittest.TestCase):
    def test_b_missing_start_balance_no_partial_sum(self):
        # 裁定: 相続開始時残高が空の行があれば集計不能＝部分合計を出さない
        records = _sample_records()
        records[3]["相続開始時残高"] = {"value": ""}
        with patch.dict(os.environ, _ENV):
            data = zx.build_zaisan_xlsx(records)
        ws = load_workbook(io.BytesIO(data)).active
        self.assertIsNone(ws["I13"].value)            # B 小計は空欄
        self.assertEqual(ws["J13"].value, zx.NOTE_B)
        self.assertIsNone(ws["I25"].value)            # 総合計も出さない
        self.assertEqual(ws["J25"].value, zx.NOTE_TOTAL)
        zx.verify_zaisan_xlsx(data, records)          # 検証も整合

    def test_mochibun_out_of_grammar_blank_plus_note(self):
        records = _sample_records()
        records[0]["特定情報"] = {"value":
                                  "所在 川口市青木1丁目 / 持分 2分の1くらい"}
        with patch.dict(os.environ, _ENV):
            data = zx.build_zaisan_xlsx(records)
        ws = load_workbook(io.BytesIO(data)).active
        self.assertIsNone(ws["I6"].value)             # 持分評価格は空欄
        self.assertIn(zx.NOTE_MOCHIBUN, ws["J6"].value)
        self.assertIsNone(ws["I8"].value)             # A 小計は算定しない
        self.assertEqual(ws["J8"].value, zx.NOTE_A)
        zx.verify_zaisan_xlsx(data, records)

    def test_zero_records_rejected(self):
        with self.assertRaises(ZaisanMokurokuError):
            zx.build_zaisan_xlsx([])

    def test_unconfirmed_valuation_builds_draft(self):
        # ZAISAN-GEN-2 裁定由来の期待値変更: 評価確定の全件要求は生成条件から
        # 外れ、未確定行があっても「下書き」（バナー+行注記+暫定表示）として
        # 生成される（拒否しない）。詳細検証は test_zaisan_gen2.py
        records = _sample_records()
        records[2]["評価確定"] = {"value": "no"}
        with patch.dict(os.environ, _ENV):
            data = zx.build_zaisan_xlsx(records)
        ws = load_workbook(io.BytesIO(data)).active
        self.assertEqual(ws["B1"].value, zx.DRAFT_BANNER)
        zx.verify_zaisan_xlsx(data, records)

    def test_template_hash_mismatch_rejected(self):
        with patch.object(zx, "TEMPLATE_SHA256", "0" * 64):
            with self.assertRaises(zx.ZaisanXlsxIntegrityError):
                zx.build_zaisan_xlsx(_sample_records())


class TestVerifyNegatives(unittest.TestCase):
    def _tampered(self, mutate):
        with patch.dict(os.environ, _ENV):
            data = zx.build_zaisan_xlsx(_sample_records())
        wb = load_workbook(io.BytesIO(data))
        mutate(wb.active)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_tampered_subtotal_rejected(self):
        data = self._tampered(lambda ws: ws.__setitem__("I13", 400001))
        with self.assertRaises(zx.ZaisanXlsxIntegrityError):
            zx.verify_zaisan_xlsx(data, _sample_records())

    def test_tampered_total_rejected(self):
        data = self._tampered(lambda ws: ws.__setitem__("I25", 3050001))
        with self.assertRaises(zx.ZaisanXlsxIntegrityError):
            zx.verify_zaisan_xlsx(data, _sample_records())

    def test_string_amount_rejected(self):
        data = self._tampered(lambda ws: ws.__setitem__("I16", "400,000"))
        with self.assertRaises(zx.ZaisanXlsxIntegrityError):
            zx.verify_zaisan_xlsx(data, _sample_records())

    def test_formula_rejected(self):
        data = self._tampered(lambda ws: ws.__setitem__("I13", "=SUM(H11:H12)"))
        with self.assertRaises(zx.ZaisanXlsxIntegrityError):
            zx.verify_zaisan_xlsx(data, _sample_records())

    def test_row_count_mismatch_rejected(self):
        with patch.dict(os.environ, _ENV):
            data = zx.build_zaisan_xlsx(_sample_records())
        with self.assertRaises(zx.ZaisanXlsxIntegrityError):
            zx.verify_zaisan_xlsx(data, _sample_records()[:-1])

    def test_consistent_simultaneous_tamper_rejected(self):
        # fix1[02]: 明細+小計+総合計を整合的に同時改変（内部整合は保たれる）
        # しても、元 records の閉集合投影との照合で拒否される
        def mutate(ws):
            ws["I16"] = 400001
            ws["I17"] = 400001
            ws["I25"] = 3050002
        data = self._tampered(mutate)
        with self.assertRaises(zx.ZaisanXlsxIntegrityError):
            zx.verify_zaisan_xlsx(data, _sample_records())

    def test_text_cell_tamper_rejected(self):
        # fix1[02]: 金額以外（所在等）の改変も対レコード照合で拒否
        data = self._tampered(lambda ws: ws.__setitem__("C6", "別の所在"))
        with self.assertRaises(zx.ZaisanXlsxIntegrityError):
            zx.verify_zaisan_xlsx(data, _sample_records())

    def test_merge_tamper_rejected(self):
        # fix1[04]: 結合レンジの欠落は完全一致検査で拒否
        data = self._tampered(lambda ws: ws.unmerge_cells("E11:G11"))
        with self.assertRaises(zx.ZaisanXlsxIntegrityError):
            zx.verify_zaisan_xlsx(data, _sample_records())


class TestLayoutVariants(unittest.TestCase):
    """fix1[04]: 可変行数の各配置で結合レンジ完全一致+書式保持。"""

    def _build(self, records):
        with patch.dict(os.environ, _ENV):
            return zx.build_zaisan_xlsx(records)

    def test_merges_exact_on_happy(self):
        ws = load_workbook(io.BytesIO(self._build(_sample_records()))).active
        self.assertEqual({str(r) for r in ws.merged_cells.ranges},
                         set(zx._merge_ranges(2, 2, 1, 1)))

    def test_empty_sections_single_row_layout(self):
        # B 部 1 件のみ（A/C/D は 0 件=空 1 行・小計 0・総合計=B）
        records = [_zrec("1", 財産種別="預貯金",
                         特定情報="テスト銀行 青木支店 普通預金 "
                                  "口座番号1234567",
                         相続開始時残高="300000")]
        data = self._build(records)
        zx.verify_zaisan_xlsx(data, records)
        ws = load_workbook(io.BytesIO(data)).active
        self.assertEqual(ws.max_row, 23)
        self.assertEqual({str(r) for r in ws.merged_cells.ranges},
                         set(zx._merge_ranges(1, 1, 1, 1)))
        self.assertEqual(ws["I7"].value, 0)       # A 小計（0 件）
        self.assertEqual(ws["I11"].value, 300000)
        self.assertEqual(ws["I15"].value, 0)
        self.assertEqual(ws["I20"].value, 0)
        self.assertEqual(ws["I23"].value, 300000)
        for col in range(2, 12):                  # 空 A プロト行は完全空欄
            self.assertIsNone(ws.cell(row=6, column=col).value)

    def test_many_rows_layout_and_style(self):
        records = (
            [_zrec(str(i), 財産種別="預貯金",
                   特定情報=f"テスト銀行 支店{i} 普通貯金 口座番号{1000+i}",
                   相続開始時残高="1000") for i in range(12)]
            + [_zrec(str(20 + i), 財産種別="葬儀費用",
                     特定情報="テスト葬祭", 評価額="500") for i in range(4)])
        data = self._build(records)
        zx.verify_zaisan_xlsx(data, records)
        ws = load_workbook(io.BytesIO(data)).active
        self.assertEqual(ws.max_row, 23 + 11 + 3)  # B+11・D+3
        self.assertEqual({str(r) for r in ws.merged_cells.ranges},
                         set(zx._merge_ranges(1, 12, 1, 4)))
        self.assertEqual(ws["I22"].value, 12000)   # B 小計（11+11）
        self.assertEqual(ws["I25"].value, None)    # C 小計行ではない位置検査
        self.assertEqual(ws.cell(row=22, column=9).value, 12000)
        self.assertEqual(ws["I37"].value, 0 + 12000 + 0 - 2000)  # 総合計
        # 挿入行の書式保持: 金額書式・行高がプロト行と同一
        proto_h = ws.row_dimensions[10].height
        for r in (11, 15, 21):
            self.assertIn("#,##0", ws.cell(row=r, column=8).number_format)
            self.assertEqual(ws.row_dimensions[r].height, proto_h)


class _WebhookBase(unittest.TestCase):
    def _post(self, *, record, zaisan=None, body=None, url=_URL,
              upload=None, update=None):
        upload = upload or AsyncMock(return_value="fk-1")
        update = update or AsyncMock()
        notify = AsyncMock(return_value=True)
        get = AsyncMock(return_value=record)
        fetch = AsyncMock(
            return_value=_sample_records() if zaisan is None else zaisan)
        with patch.dict(os.environ, _ENV), \
             patch.object(zw.hub_kintone, "get_record", get), \
             patch.object(zw.hub_kintone, "upload_file", upload), \
             patch.object(zw.hub_kintone, "update_record", update), \
             patch.object(zw, "fetch_zaisan_records", fetch), \
             patch("hub.notify.notify_admin_line", notify):
            r = _client.post(url, json=body or _body())
        return r, upload, update, notify, get


def _body(record_id="7", status="財産目録作成", app_id="26"):
    body = {"record": {"$id": {"value": record_id},
                       "財産目録ステータス": {"value": status}}}
    if app_id is not None:
        body["app"] = {"id": app_id}
    return body


def _case_record(**over):
    base = {"$revision": "5", "財産目録ステータス": "財産目録作成",
            "被相続人名": "試験太郎", "財産目録": []}
    base.update(over)
    return {k: {"value": v} for k, v in base.items()}


class TestWebhookEntry(_WebhookBase):
    def test_wrong_secret_403(self):
        r, *_ = self._post(record=_case_record(), url="/zaisan/wrong")
        self.assertEqual(r.status_code, 403)

    def test_app_mismatch_zero_effects(self):
        for app_id in ("21", None, "abc"):
            with self.subTest(app_id=app_id):
                r, upload, update, _n, get = self._post(
                    record=_case_record(), body=_body(app_id=app_id))
                self.assertEqual(r.json().get("skip"), "app_mismatch")
                get.assert_not_awaited()
                upload.assert_not_awaited()

    def test_not_triggered_body_gate(self):
        for status in ("財産目録作成中", "財産目録作成済", "要確認", ""):
            with self.subTest(status=status):
                r, upload, _u, _n, get = self._post(
                    record=_case_record(), body=_body(status=status))
                self.assertEqual(r.json().get("skip"), "not_triggered")
                get.assert_not_awaited()
                upload.assert_not_awaited()

    def test_stale_authoritative_states(self):
        cases = {"財産目録作成済": "already_done", "要確認": "stale_status",
                 "": "stale_status"}
        for state, skip in cases.items():
            with self.subTest(state=state):
                r, upload, update, _n, _g = self._post(
                    record=_case_record(財産目録ステータス=state))
                self.assertEqual(r.json().get("skip"), skip)
                upload.assert_not_awaited()
                update.assert_not_awaited()


class TestWebhookStateMachine(_WebhookBase):
    def test_happy_path(self):
        r, upload, update, notify, _g = self._post(record=_case_record())
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.json().get("record_id"), "7")
        self.assertEqual(update.await_count, 2)
        cas_call, final_call = update.await_args_list
        self.assertEqual(cas_call.args[2],
                         {"財産目録ステータス": "財産目録作成中"})
        self.assertEqual(cas_call.kwargs.get("revision"), "5")
        self.assertEqual(final_call.args[2]["財産目録ステータス"],
                         "財産目録作成済")
        self.assertEqual(final_call.args[2]["財産目録"],
                         [{"fileKey": "fk-1"}])
        self.assertEqual(final_call.kwargs.get("revision"), "6")
        # 添付物は検証済み実体（xlsx・被相続人名差し込み・総合計一致）
        fname, data, mime = (upload.await_args.args[1],
                             upload.await_args.args[2],
                             upload.await_args.args[3])
        self.assertEqual(fname, "財産目録.xlsx")
        self.assertIn("spreadsheetml", mime)
        zx.verify_zaisan_xlsx(data, _sample_records())
        ws = load_workbook(io.BytesIO(data)).active
        self.assertEqual(ws["C3"].value, "被相続人　試験太郎")
        # ZAISAN-GEN-2 裁定由来: 生成通知（完成版の明記）が 1 回入る
        notify.assert_awaited_once()
        self.assertIn("完成版として生成", notify.await_args.args[0])

    def test_cas_loser_zero_effects(self):
        loser = AsyncMock(side_effect=KintoneError(409, "GAIA_CO02"))
        r, upload, update, _n, _g = self._post(
            record=_case_record(), update=loser)
        self.assertEqual(r.json().get("skip"), "cas_lost")
        upload.assert_not_awaited()
        self.assertEqual(update.await_count, 1)

    def test_claim_failures_not_silenced(self):
        for err in (KintoneError(500, "GAIA_XX01", "x"),
                    KintoneError(0, "transport_error", "x")):
            with self.subTest(status=err.status):
                update = AsyncMock(side_effect=err)
                r, upload, _u, _n, _g = self._post(
                    record=_case_record(), update=update)
                self.assertEqual(r.status_code, 500)
                upload.assert_not_awaited()

    def test_recovery_no_attachment(self):
        record = _case_record(財産目録ステータス="財産目録作成中")
        record["$revision"] = {"value": "7"}
        r, upload, update, notify, _g = self._post(record=record)
        self.assertTrue(r.json().get("recovered"))
        self.assertEqual(upload.await_count, 1)
        self.assertEqual(update.await_args_list[1].kwargs.get("revision"),
                         "8")
        # ZAISAN-GEN-2 裁定由来: 回収成功時も生成通知（完成版）が入る
        notify.assert_awaited_once()
        self.assertIn("完成版として生成", notify.await_args.args[0])

    def test_working_with_attachment_goes_review(self):
        record = _case_record(財産目録ステータス="財産目録作成中")
        record["財産目録"] = {"value": [{"fileKey": "old"}]}
        r, upload, update, notify, _g = self._post(record=record)
        self.assertEqual(r.json().get("skip"), "needs_review")
        upload.assert_not_awaited()
        self.assertEqual(update.await_args.args[2],
                         {"財産目録ステータス": "要確認"})
        notify.assert_awaited_once()
        self.assertIn("自動では上書きせず", notify.await_args.args[0])

    def test_upload_failure_500_stays_working(self):
        upload = AsyncMock(side_effect=RuntimeError("kintone down"))
        r, _u, update, _n, _g = self._post(record=_case_record(),
                                           upload=upload)
        self.assertEqual(r.status_code, 500)
        self.assertEqual(update.await_count, 1)  # CAS のみ


class TestWebhookFailClosed(_WebhookBase):
    def test_zero_zaisan_records_rejected(self):
        r, upload, update, notify, _g = self._post(
            record=_case_record(), zaisan=[])
        self.assertEqual(r.json().get("skip"), "not_ready")
        upload.assert_not_awaited()
        update.assert_not_awaited()              # 状態も動かさない
        notify.assert_awaited_once()
        sent = notify.await_args.args[0]
        self.assertIn("財産行が0件", sent)
        self.assertNotIn("試験", sent)           # PII 非搭載

    def test_unconfirmed_generates_draft_with_notice(self):
        # ZAISAN-GEN-2 裁定由来の期待値変更: 評価未確定は拒否せず下書きとして
        # 生成し、通知に「下書き（評価未確定 N 件）」を明記する
        records = _sample_records()
        records[0]["評価確定"] = {"value": "no"}
        r, upload, update, notify, _g = self._post(
            record=_case_record(), zaisan=records)
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.json().get("record_id"), "7")
        upload.assert_awaited_once()
        zx.verify_zaisan_xlsx(upload.await_args.args[2], records)
        notify.assert_awaited_once()
        sent = notify.await_args.args[0]
        self.assertIn("下書き（評価未確定 1 件）", sent)
        self.assertNotIn("試験", sent)           # PII 非搭載


class TestSchemaPin(unittest.TestCase):
    def test_app26_fields(self):
        fields = EXPECTED_KINTONE_SCHEMA["相談カード (相続)"]["fields"]
        self.assertEqual(fields["財産目録ステータス"]["required_options"],
                         ["財産目録作成", "財産目録作成中",
                          "財産目録作成済", "要確認"])
        self.assertEqual(fields["財産目録"]["type"], "FILE")

    def test_app35_fields(self):
        fields = EXPECTED_KINTONE_SCHEMA["App 35 (財産)"]["fields"]
        self.assertEqual(fields["資料番号"]["type"], "SINGLE_LINE_TEXT")
        self.assertEqual(fields["相続開始時残高"]["type"], "NUMBER")
        self.assertEqual(fields["現在残高"]["type"], "NUMBER")


if __name__ == "__main__":
    unittest.main()
