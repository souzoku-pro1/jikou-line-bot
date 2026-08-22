"""ZAISAN-GEN-2: 財産目録の下書き生成（大野裁定）。

固定する仕様:
- 生成条件は「有効な財産レコード 1 件以上」のみ（評価確定の全件要求を
  生成条件から外す）。0 件拒否は不変。
- 「機械は確定しない」は維持: 評価確定≠yes の行は備考へ「※評価未確定」を
  自動追記（既存注記とは「／」連結）。
- 未確定行が 1 つでもあれば: 表題直上の余白行（B1:K1 結合・原本の空行）へ
  下書きバナー+算定できた小計・総合計の注記欄（J 列）へ「（暫定）」。
  金額セルは int のまま（暫定表示は注記欄＝検証水準を落とさない）。
- 全件確定+全金額入力は従来どおり注記なしの完成版（バナー・暫定なし）。
- verify は下書きモードも対レコード照合（未確定注記・バナー・暫定表示を
  閉集合として照合・完成版への混入も両方向で拒否）。
- 生成通知に「下書き（評価未確定 N 件）」/「完成版」を明記。
"""

import io
import os
import unittest
from unittest.mock import patch

from test_zaisan_gen1 import (  # noqa: F401
    _ENV, _WebhookBase, _body, _case_record, _sample_records, _zrec)

from openpyxl import load_workbook  # noqa: E402

from units.souzoku import zaisan_xlsx as zx  # noqa: E402


def _draft_records():
    records = _sample_records()
    records[2]["評価確定"] = {"value": "no"}     # B 部 1 行を未確定に
    return records


class TestDraftBuild(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        with patch.dict(os.environ, _ENV):
            cls.data = zx.build_zaisan_xlsx(_draft_records())
        cls.ws = load_workbook(io.BytesIO(cls.data)).active

    def test_banner_on_spacer_row(self):
        # 表題直上の余白行（B1:K1 結合）にバナー＝テンプレ体裁を崩さない配置
        self.assertEqual(self.ws["B1"].value, zx.DRAFT_BANNER)
        self.assertEqual(self.ws["B2"].value.strip(),
                         "遺　　産　　目　　録")   # 表題は不変

    def test_unconfirmed_row_note_appended(self):
        # 未確定行（B 部 1 行目）の備考に「※評価未確定」
        self.assertEqual(self.ws["J11"].value, zx.NOTE_UNCONFIRMED)
        # 確定行には付かない
        self.assertIsNone(self.ws["J12"].value)

    def test_note_joined_with_existing_note(self):
        records = _draft_records()
        records[5]["評価確定"] = {"value": "no"}  # 備考「立替済」の行
        with patch.dict(os.environ, _ENV):
            data = zx.build_zaisan_xlsx(records)
        ws = load_workbook(io.BytesIO(data)).active
        self.assertEqual(ws["J21"].value,
                         f"立替済／{zx.NOTE_UNCONFIRMED}")
        zx.verify_zaisan_xlsx(data, records)

    def test_provisional_marks_and_int_amounts(self):
        # 算定できた小計・総合計は int のまま+J 列に「（暫定）」
        for sub_cell, note_cell, want in (("I8", "J8", 2500000),
                                          ("I13", "J13", 400000),
                                          ("I17", "J17", 400000),
                                          ("I22", "J22", 250000),
                                          ("I25", "J25", 3050000)):
            self.assertEqual(self.ws[sub_cell].value, want)
            self.assertEqual(self.ws[note_cell].value, zx.PROVISIONAL_MARK)

    def test_verify_passes_draft(self):
        zx.verify_zaisan_xlsx(self.data, _draft_records())

    def test_final_version_unchanged(self):
        # 全件確定は従来どおり注記なしの完成版（バナー・暫定なし）
        with patch.dict(os.environ, _ENV):
            data = zx.build_zaisan_xlsx(_sample_records())
        ws = load_workbook(io.BytesIO(data)).active
        self.assertIsNone(ws["B1"].value)
        for cell in ("J8", "J13", "J17", "J25"):
            self.assertIsNone(ws[cell].value)
        zx.verify_zaisan_xlsx(data, _sample_records())

    def test_draft_with_uncomputable_section_keeps_note(self):
        # 未確定+B 部の相続開始時残高欠落: 集計不能注記が優先（暫定は
        # 算定できた部のみ）・総合計は算定不能注記
        records = _draft_records()
        records[3]["相続開始時残高"] = {"value": ""}
        with patch.dict(os.environ, _ENV):
            data = zx.build_zaisan_xlsx(records)
        ws = load_workbook(io.BytesIO(data)).active
        self.assertIsNone(ws["I13"].value)
        self.assertEqual(ws["J13"].value, zx.NOTE_B)
        self.assertEqual(ws["J8"].value, zx.PROVISIONAL_MARK)  # A は暫定
        self.assertIsNone(ws["I25"].value)
        self.assertEqual(ws["J25"].value, zx.NOTE_TOTAL)
        zx.verify_zaisan_xlsx(data, records)

    def test_zero_records_still_rejected(self):
        from units.souzoku.zaisan_mokuroku import ZaisanMokurokuError
        with self.assertRaises(ZaisanMokurokuError):
            zx.build_zaisan_xlsx([])


class TestDraftVerifyNegatives(unittest.TestCase):
    def _tampered(self, records, mutate):
        with patch.dict(os.environ, _ENV):
            data = zx.build_zaisan_xlsx(records)
        wb = load_workbook(io.BytesIO(data))
        mutate(wb.active)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_missing_banner_rejected(self):
        data = self._tampered(_draft_records(),
                              lambda ws: ws.__setitem__("B1", None))
        with self.assertRaises(zx.ZaisanXlsxIntegrityError):
            zx.verify_zaisan_xlsx(data, _draft_records())

    def test_banner_on_final_rejected(self):
        data = self._tampered(_sample_records(),
                              lambda ws: ws.__setitem__("B1",
                                                        zx.DRAFT_BANNER))
        with self.assertRaises(zx.ZaisanXlsxIntegrityError):
            zx.verify_zaisan_xlsx(data, _sample_records())

    def test_missing_provisional_mark_rejected(self):
        data = self._tampered(_draft_records(),
                              lambda ws: ws.__setitem__("J25", None))
        with self.assertRaises(zx.ZaisanXlsxIntegrityError):
            zx.verify_zaisan_xlsx(data, _draft_records())

    def test_missing_unconfirmed_note_rejected(self):
        data = self._tampered(_draft_records(),
                              lambda ws: ws.__setitem__("J11", None))
        with self.assertRaises(zx.ZaisanXlsxIntegrityError):
            zx.verify_zaisan_xlsx(data, _draft_records())

    def test_provisional_on_final_rejected(self):
        data = self._tampered(_sample_records(),
                              lambda ws: ws.__setitem__(
                                  "J25", zx.PROVISIONAL_MARK))
        with self.assertRaises(zx.ZaisanXlsxIntegrityError):
            zx.verify_zaisan_xlsx(data, _sample_records())


class TestWebhookDraftNotice(_WebhookBase):
    def test_draft_notice_counts(self):
        records = _draft_records()
        records[3]["評価確定"] = {"value": "no"}
        r, upload, _u, notify, _g = self._post(
            record=_case_record(), zaisan=records)
        self.assertEqual(r.status_code, 200)
        upload.assert_awaited_once()
        notify.assert_awaited_once()
        self.assertIn("下書き（評価未確定 2 件）", notify.await_args.args[0])

    def test_final_notice(self):
        r, upload, _u, notify, _g = self._post(record=_case_record())
        upload.assert_awaited_once()
        notify.assert_awaited_once()
        self.assertIn("完成版として生成", notify.await_args.args[0])


if __name__ == "__main__":
    unittest.main()
